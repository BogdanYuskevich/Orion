from django.views.generic import ListView, UpdateView, DeleteView, DetailView
from django.urls import reverse_lazy
from django.shortcuts import render, redirect
from .models import Order, OrderItem
from .forms import OrderForm, OrderItemFormSet
from notifications.order_email_notifications import send_order_notification  # новий модуль
from django.db.models import Q
from django.views.generic import TemplateView
from django.db.models import Sum, Count

import csv
import io
from django.http import HttpResponse
from django.views import View

import openpyxl
from openpyxl.utils import get_column_letter
from django.views.generic import ListView
from .models import Order

class OrderListView(ListView):
    model = Order
    template_name = 'orders/order_list.html'
    context_object_name = 'orders'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset()
        # Пошуковий запит: клієнт або пошук по статусу (як підрядок)
        search_query = self.request.GET.get("q")
        if search_query:
            queryset = queryset.filter(
                Q(customer__first_name__icontains=search_query) |
                Q(status__icontains=search_query)
            )
        # Фільтрація по статусу із select (якщо вибрано конкретний статус)
        status_filter = self.request.GET.get("status")
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        # Сортування
        sort_by = self.request.GET.get("sort_by")
        order = self.request.GET.get("order", "asc")
        if sort_by:
            if order == "desc":
                sort_by = "-" + sort_by
            queryset = queryset.order_by(sort_by)
        else:
            queryset = queryset.order_by("-order_date")
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Передаємо в контекст пошукове значення
        context['q'] = self.request.GET.get("q", "")
        # Передаємо поточний відфільтрований статус
        context['status_filter'] = self.request.GET.get("status", "")
        # Якщо у моделі Order визначено choices для поля status – використаємо їх,
        # інакше візьмемо всі відмінні значення статусу з БД
        try:
            context['status_choices'] = Order._meta.get_field('status').choices
        except Exception:
            context['status_choices'] = Order.objects.order_by('status').values_list('status', flat=True).distinct()
        return context

    def render_to_response(self, context, **response_kwargs):
        if self.request.GET.get("export") == "csv":
            queryset = self.get_queryset()
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = 'attachment; filename="orders.csv"'
            writer = csv.writer(response)
            writer.writerow(["ID", "Клієнт", "Дата замовлення", "Статус", "Сума"])
            for order in queryset:
                writer.writerow([
                    order.id,
                    order.customer,
                    order.order_date.strftime("%Y-%m-%d %H:%M"),
                    order.status,
                    order.total_amount,
                ])
            return response
        else:
            return super().render_to_response(context, **response_kwargs)


def order_create(request):
    if request.method == 'POST':
        form = OrderForm(request.POST)
        # Використовуємо instance=Order() для створення нового замовлення
        formset = OrderItemFormSet(request.POST, instance=Order())
        if form.is_valid() and formset.is_valid():
            order = form.save(commit=False)
            order.total_amount = 0  # Розрахунок загальної суми пізніше
            order.save()

            total = 0
            formset.instance = order
            items = formset.save(commit=False)
            for item in items:
                # Зафіксуємо ціну продукту на момент замовлення
                item.price_at_order = item.product.price
                total += item.product.price * item.quantity
                item.save()
            order.total_amount = total
            order.save()

            # Надсилаємо повідомлення про нове замовлення із переліком усіх позицій
            send_order_notification(
                receiver_email="yuskevich12@gmail.com",  # адреса, куди надсилати повідомлення
                order=order,
                order_items=items
            )

            return redirect('order-list')
    else:
        form = OrderForm()
        formset = OrderItemFormSet(instance=Order(), queryset=OrderItem.objects.none())
    return render(request, 'orders/order_form.html', {'form': form, 'formset': formset})




class ExportOrdersXLSX(View):
    def get(self, request, *args, **kwargs):
        orders = Order.objects.all()  # або з get_queryset() для фільтрації
        output = io.BytesIO()

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Orders'

        # Заголовки
        headers = ["ID", "Клієнт", "Дата замовлення", "Статус", "Сума"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            worksheet[f"{col_letter}1"] = header

        for row_num, order in enumerate(orders, start=2):
            worksheet[f"A{row_num}"] = order.id
            worksheet[f"B{row_num}"] = str(order.customer)
            worksheet[f"C{row_num}"] = order.order_date.strftime("%Y-%m-%d %H:%M")
            worksheet[f"D{row_num}"] = order.status
            worksheet[f"E{row_num}"] = order.total_amount

        workbook.save(output)
        output.seek(0)
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=orders.xlsx'
        return response



class OrderAnalyticsView(TemplateView):
    template_name = 'orders/order_analytics.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        orders = Order.objects.all()
        total_orders = orders.count()
        total_revenue = orders.aggregate(total=Sum('total_amount'))['total'] or 0

        # Групуємо замовлення за статусом для побудови діаграми
        status_data = orders.values('status').annotate(count=Count('id')).order_by('status')
        labels = [item['status'] for item in status_data]
        data = [item['count'] for item in status_data]

        context.update({
            'total_orders': total_orders,
            'total_revenue': total_revenue,
            'labels': labels,
            'data': data,
        })
        return context

class OrderUpdateView(UpdateView):
    model = Order
    form_class = OrderForm
    template_name = 'orders/order_form.html'
    success_url = reverse_lazy('order-list')

class OrderDeleteView(DeleteView):
    model = Order
    template_name = 'orders/order_confirm_delete.html'
    success_url = reverse_lazy('order-list')

class OrderDetailView(DetailView):
    model = Order
    template_name = 'orders/order_detail.html'
