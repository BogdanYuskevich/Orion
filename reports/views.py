from django.shortcuts import render
from django.http import HttpResponse
from io import BytesIO
from datetime import datetime, timedelta
from django.utils import timezone
from orders.models import Order
from django.db.models import Sum, Avg

# Для експорту в PDF:
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os
from django.conf import settings

# Для експорту в Excel
import openpyxl
from openpyxl.styles import Font

def generate_report_data():
    """
    Генерує детальні дані звіту...
    """
    today = timezone.now().date()
    yesterday = today - timedelta(days=1)

    orders_today_qs = Order.objects.filter(order_date__date=today)
    today_count = orders_today_qs.count()
    revenue_today = orders_today_qs.aggregate(total=Sum('total_amount'))['total'] or 0
    avg_order = orders_today_qs.aggregate(avg=Avg('total_amount'))['avg'] or 0

    orders_yesterday_qs = Order.objects.filter(order_date__date=yesterday)
    yesterday_count = orders_yesterday_qs.count()

    if yesterday_count > 0:
        change = ((today_count - yesterday_count) / yesterday_count * 100)
    else:
        change = 0

    orders_list = []
    for order in orders_today_qs:
        orders_list.append({
            'id': order.id,
            'date': order.order_date.strftime("%Y-%m-%d %H:%M"),
            'amount': float(order.total_amount),
        })

    chart_labels = []
    chart_data = []
    for i in range(7):
        day = today - timedelta(days=6 - i)
        label = day.strftime("%Y-%m-%d")
        count = Order.objects.filter(order_date__date=day).count()
        chart_labels.append(label)
        chart_data.append(count)

    return {
        'today': today,
        'today_count': today_count,
        'yesterday_count': yesterday_count,
        'change_percentage': round(change, 2),
        'change_percentage_abs': round(abs(change), 2),  # додано
        'revenue_today': round(revenue_today, 2),
        'avg_order': round(avg_order, 2),
        'orders_list': orders_list,
        'chart_labels': chart_labels,
        'chart_data': chart_data,
    }


def daily_report_view(request):
    report_data = generate_report_data()
    return render(request, "reports/daily_report.html", report_data)


def export_pdf_report(request):
    """
    Генерація PDF‑звіту з використанням TTF‑шрифту (DejaVuSans.ttf) для коректного відображення українських символів.
    """
    buffer = BytesIO()
    p = canvas.Canvas(buffer)

    font_path = os.path.join(settings.BASE_DIR, 'DejaVuSans.ttf')
    try:
        pdfmetrics.registerFont(TTFont('DejaVu', font_path))
        font_name = 'DejaVu'
    except Exception as e:
        font_name = 'Helvetica'

    report_data = generate_report_data()
    report_date = report_data['today'].strftime("%Y-%m-%d")
    today_count = report_data['today_count']
    yesterday_count = report_data['yesterday_count']
    change = report_data['change_percentage']
    revenue = report_data['revenue_today']
    avg_order = report_data['avg_order']

    y_position = 800
    p.setFont(font_name, 16)
    p.drawString(100, y_position, f"Щоденний звіт за {report_date}")
    y_position -= 30
    p.setFont(font_name, 12)
    p.drawString(100, y_position, f"Замовлень сьогодні: {today_count}")
    y_position -= 20
    p.drawString(100, y_position, f"Замовлень вчора: {yesterday_count}")
    y_position -= 20
    p.drawString(100, y_position, f"Зміна: {change}%")
    y_position -= 20
    p.drawString(100, y_position, f"Загальна сума сьогодні: {revenue}")
    y_position -= 20
    p.drawString(100, y_position, f"Середня сума замовлення: {avg_order}")

    # Додамо невелкий розбір списку замовлень (не обов'язково, але для прикладу)
    y_position -= 40
    p.setFont(font_name, 12)
    p.drawString(100, y_position, "Деталі замовлень:")
    y_position -= 20
    for order in report_data['orders_list']:
        # Якщо місце обмежене, можна виводити тільки перші кілька
        p.drawString(110, y_position, f"№{order['id']}, {order['date']}, сума: {order['amount']}")
        y_position -= 15
        if y_position < 100:
            p.showPage()
            y_position = 800
            p.setFont(font_name, 12)

    p.showPage()
    p.save()
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf')


def export_excel_report(request):
    report_data = generate_report_data()
    report_date = report_data['today'].strftime("%Y-%m-%d")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Report"

    header_font = Font(bold=True)
    ws['A1'] = "Дата звіту"
    ws['B1'] = report_date

    ws['A2'] = "Замовлень сьогодні"
    ws['A2'].font = header_font
    ws['B2'] = report_data['today_count']

    ws['A3'] = "Замовлень вчора"
    ws['A3'].font = header_font
    ws['B3'] = report_data['yesterday_count']

    ws['A4'] = "Зміна (%)"
    ws['A4'].font = header_font
    ws['B4'] = report_data['change_percentage']

    ws['A5'] = "Загальна сума сьогодні"
    ws['A5'].font = header_font
    ws['B5'] = report_data['revenue_today']

    ws['A6'] = "Середня сума замовлення"
    ws['A6'].font = header_font
    ws['B6'] = report_data['avg_order']

    # Додамо таблицю з деталями замовлень
    ws.cell(row=8, column=1, value="№ Замовлення").font = header_font
    ws.cell(row=8, column=2, value="Дата").font = header_font
    ws.cell(row=8, column=3, value="Сума").font = header_font

    row = 9
    for order in report_data['orders_list']:
        ws.cell(row=row, column=1, value=order['id'])
        ws.cell(row=row, column=2, value=order['date'])
        ws.cell(row=row, column=3, value=order['amount'])
        row += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"daily_report_{report_date}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
